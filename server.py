from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from ortools.sat.python import cp_model
from typing import List, Optional, Dict
import pandas as pd
import io
from fastapi.responses import Response, HTMLResponse, FileResponse
import uuid
import datetime
import calendar
import json
import os
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi import Depends
from starlette.status import HTTP_401_UNAUTHORIZED
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

app = FastAPI()
security = HTTPBasic()

def authenticate(credentials: HTTPBasicCredentials = Depends(security)):
    correct_password = os.getenv("AUTH_PASSWORD", "admin") # Default to admin if not set
    if credentials.username != "admin" or credentials.password != correct_password:
        raise HTTPException(
            status_code=HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username

# Ensure storage directories exist
os.makedirs("schedules", exist_ok=True)
os.makedirs("static", exist_ok=True)

# Cleanup old temp export files on startup
try:
    for f in os.listdir("schedules"):
        if f.startswith("export_") and f.endswith(".xlsx"):
            os.remove(os.path.join("schedules", f))
    print("Cleaned up temporary export files.")
except Exception as e:
    print(f"Cleanup error: {e}")

# ---- Pydantic Models ----
class Staff(BaseModel):
    id: str
    name: str
    can_night: bool
    floor: int  # 1 or 2
    restricted_shift: Optional[str] = None # e.g. "早③"
    is_weekday_only: bool = False

class ScheduleRequest(BaseModel):
    year: int = datetime.datetime.now().year
    month: int = datetime.datetime.now().month
    staff_list: List[Staff]
    # Previous month's last 3 shifts for each staff (name -> [shift1, shift2, shift3])
    previous_shifts: Dict[str, List[str]] = {}
    previous_holiday_counts: Dict[str, float] = {}
    # Current table state to fix specific cells (name -> [shift1, shift2, ... or null])
    current_shifts: Optional[List[dict]] = None

# ---- Scheduling Logic ----
def generate_schedule(year: int, month: int, staff_list: List[Staff], prev_shifts_map: Dict[str, List[str]], prev_holiday_counts: Dict[str, float], fixed_shifts_data: Optional[List[dict]] = None):
    _, num_days = calendar.monthrange(year, month)
    num_staff = len(staff_list)
    # Shift types: 0: 休, 1: 早②, 2: 早③, 3: 日, 4: 遅①, 5: 遅②, 6: 夜①, 7: 夜②, 8: 明け, 9: 病休, 10: 欠勤
    num_shifts = 11
    shift_names = ["休", "早②", "早③", "日", "遅①", "遅②", "夜①", "夜②", "明け", "病休", "欠勤"]
    
    model = cp_model.CpModel()
    x = {}
    for e in range(num_staff):
        for d in range(num_days):
            for s in range(num_shifts):
                x[(e, d, s)] = model.NewBoolVar(f'shift_e{e}d{d}s{s}')
                
    # 1. Exactly one shift per day
    for e in range(num_staff):
        for d in range(num_days):
            model.AddExactlyOne(x[(e, d, s)] for s in range(num_shifts))

    # --- 2.5 Parse Fixed (Manual) Shifts (Ver 8.23: Moved to top) ---
    fixed_indices = {} # (e, d) -> shift_index
    if fixed_shifts_data:
        fixed_map = {item['staff_id']: item['shifts'] for item in fixed_shifts_data}
        for e, staff in enumerate(staff_list):
            if staff.name in fixed_map:
                f_shifts = fixed_map[staff.name]
                for d in range(min(num_days, len(f_shifts))):
                    fs = f_shifts[d]
                    if fs and fs in shift_names:
                        s_idx = shift_names.index(fs)
                        fixed_indices[(e, d)] = s_idx
    
    # Identify Absence Categories (Ver 8.21, 8.23)
    long_term_absent = []
    is_full_absent = []
    for e in range(num_staff):
        sick_count = sum(1 for d in range(num_days) if fixed_indices.get((e, d)) in [9, 10])
        long_term_absent.append(sick_count >= 7)
        is_full_absent.append(sick_count >= num_days) # Full Month Absent

    # 2. Daily requirement (Soft Constraints for floor coverage)
    mandatory_shift_vars = []
    for d in range(num_days):
        for floor_id in [1, 2]:
            for s_idx in [1, 2, 4, 5, 6, 7]:
                # Ver 8.24: Reward coverage (at least 1 person) instead of individual counts
                staff_who_can_work = [e for e, s in enumerate(staff_list) if s.floor == floor_id and not is_full_absent[e]]
                if not staff_who_can_work: continue
                
                met = model.NewBoolVar(f'd{d}_f{floor_id}_s{s_idx}_met')
                # If at least one person works this shift, 'met' is true
                model.Add(sum(x[(e, d, s_idx)] for e in staff_who_can_work) >= 1).OnlyEnforceIf(met)
                model.Add(sum(x[(e, d, s_idx)] for e in staff_who_can_work) == 0).OnlyEnforceIf(met.Not())
                mandatory_shift_vars.append((d, floor_id, s_idx, met))

    # Priority goals:
    # 1. Get as many people to 10 days (20 points) as possible.
    # 2. Prefer weekday-only staff (Managers/Business) to follow their pattern (Mon-Fri=Day, Sat-Sun=Off).
    # 3. Maximize additional Day Shifts (日勤).
    weekday_pattern_vars = []
    for e, staff in enumerate(staff_list):
        if is_full_absent[e] or long_term_absent[e]: continue 
        
        # Ver 8.25: Restrict the Mon-Fri=Day, Sat-Sun=Off pattern to Managers/Business staff only
        if staff.is_weekday_only:
            for d in range(num_days):
                # If this slot is manually fixed or an absence, skip pattern evaluation
                if fixed_indices.get((e, d)) is not None: continue 
                
                dt = datetime.date(year, month, d + 1)
                if dt.weekday() >= 5: # Sat, Sun
                    weekday_pattern_vars.append(x[(e, d, 0)]) # Goal: Off(0)
                else:
                    weekday_pattern_vars.append(x[(e, d, 3)]) # Goal: Day(3)

    s3_vars = []
    for e in range(num_staff):
        if not is_full_absent[e]:
            for d in range(num_days):
                s3_vars.append(x[(e, d, 3)])

    off_point_vars = []
    for e in range(num_staff):
        op = model.NewIntVar(0, 100, f'off_points_e{e}')
        # Holiday (0), Ake (8), Sick (9), Absent (10) contribute to off points
        # Sick and Absent count as full off days (2 points, same as '休')
        model.Add(op == sum(x[(e, d, 0)] * 2 for d in range(num_days)) + 
                         sum(x[(e, d, 8)] * 1 for d in range(num_days)) +
                         sum(x[(e, d, 9)] * 2 for d in range(num_days)) +
                         sum(x[(e, d, 10)] * 2 for d in range(num_days)))
        off_point_vars.append(op)

    is_target_holiday = []
    for e in range(num_staff):
        target_met = model.NewBoolVar(f'target_met_e{e}')
        # We define "Target Met" as having at least 20 points (10 days)
        model.Add(off_point_vars[e] >= 20).OnlyEnforceIf(target_met)
        is_target_holiday.append(target_met)

    # Priority weights (Ver 8.17):
    # 1. Mandatory Floor Coverage (Early/Late/Night): 10,000,000 (Safety & Critical Operations)
    # 2. Manager Pattern (Mon-Fri=Day, Sat-Sun=Off): 1,000,000 (Professional Standard)
    # 3. Optional Day Shifts (maximize staff usage): 500,000 (Operational Capacity)
    # 4. Staff Holiday Goals (10 days): 10,000 (Employee Perks)

    model.Maximize(
        sum(v for _, _, _, v in mandatory_shift_vars) * 10000000 + 
        sum(weekday_pattern_vars) * 1000000 + 
        sum(s3_vars) * 500000 +
        sum(is_target_holiday) * 10000
    )
            
    # 3. Night -> 明け -> 休 (Strict 3-day pattern)
    for e in range(num_staff):
        if is_full_absent[e]: continue # Ver 8.23
        for d in range(num_days - 2):
            # Rule (Ver 8.18): Night (6,7) -> Next day is (Off(0) or Ake(8) or Sick(9) or Absent(10))
            # Next next day is (Off(0) or Sick(9) or Absent(10))
            is_night = model.NewBoolVar(f'e{e}_d{d}_night')
            model.Add(sum(x[(e, d, s)] for s in [6, 7]) == 1).OnlyEnforceIf(is_night)
            model.Add(sum(x[(e, d, s)] for s in [6, 7]) == 0).OnlyEnforceIf(is_night.Not())
            
            # Day d+1: Off or Ake or Sick or Absent (Ver 8.22: added Off(0))
            model.Add(sum(x[(e, d+1, s)] for s in [0, 8, 9, 10]) == 1).OnlyEnforceIf(is_night)
            # Day d+2: Off or Sick or Absent
            model.Add(sum(x[(e, d+2, s)] for s in [0, 9, 10]) == 1).OnlyEnforceIf(is_night)

        if num_days >= 2:
            # Last day Night -> Ake
            is_night_last = model.NewBoolVar(f'e{e}_last_night')
            model.Add(sum(x[(e, num_days-2, s)] for s in [6, 7]) == 1).OnlyEnforceIf(is_night_last)
            model.Add(sum(x[(e, num_days-2, s)] for s in [6, 7]) == 0).OnlyEnforceIf(is_night_last.Not())
            model.Add(sum(x[(e, num_days-1, s)] for s in [0, 8, 9, 10]) == 1).OnlyEnforceIf(is_night_last)

        # 3.1 Month Boundary Continuity
        staff = staff_list[e]
        staff_name = staff.name.strip() # Define staff_name here
        if staff_name in prev_shifts_map:
            p_shifts = prev_shifts_map[staff_name]
            if p_shifts and not is_full_absent[e]: # Ver 8.23 skip logic
                last_shift = p_shifts[-1]
                if last_shift in ["夜①", "夜②"]:
                    # Day 0 must be Off or Ake or Sick or Absent (Ver 8.22: added Off(0))
                    model.Add(sum(x[(e, 0, s)] for s in [0, 8, 9, 10]) == 1)
                    if num_days >= 2: 
                        # Day 1 must be Off or Sick or Absent
                        model.Add(sum(x[(e, 1, s)] for s in [0, 9, 10]) == 1)
                if last_shift == "明け":
                    # If prev was Ake, Day 0 must be Off or Sick or Absent
                    model.Add(sum(x[(e, 0, s)] for s in [0, 9, 10]) == 1)
                if len(p_shifts) >= 2:
                    scnd_last = p_shifts[-2]
                    if scnd_last in ["夜①", "夜②"] and last_shift == "明け":
                        # Pattern Night -> Ake completes with Off or Sick or Absent
                        model.Add(sum(x[(e, 0, s)] for s in [0, 9, 10]) == 1)

    # (Parsing block moved up to section 2.5)
    pass

    # 4. Night Availability & Restricted Shifts
    for e, staff in enumerate(staff_list):
        if not staff.can_night:
            for d in range(num_days):
                if (e, d) in fixed_indices: continue # Human override
                model.Add(x[(e, d, 6)] == 0)
                model.Add(x[(e, d, 7)] == 0)

        if staff.restricted_shift and staff.restricted_shift in shift_names:
            s_idx = shift_names.index(staff.restricted_shift)
            for d in range(num_days):
                if (e, d) in fixed_indices: continue # Human override
                model.Add(sum(x[(e, d, s)] for s in range(num_shifts) if s not in [0, s_idx]) == 0)
        
        # 4.2 Weekday Only Handling (Moved to Soft Constraint in Objective Function)
        # We no longer Add hard constraints for (is_weekday_only) to allow manual overrides 
        # and prevent infeasibility. 
        pass

    # 4.3 Default: Prohibit AI from picking Sick/Absent (indices 9, 10) for unassigned slots
    # This ensures they only appear if manually fixed by user

    for e in range(num_staff):
        for d in range(num_days):
            # If (e, d) is not manually fixed to 9 or 10, prohibit the AI from choosing them
            if fixed_indices.get((e, d)) == 9:
                model.Add(x[(e, d, 9)] == 1)
            else:
                model.Add(x[(e, d, 9)] == 0)
                
            if fixed_indices.get((e, d)) == 10:
                model.Add(x[(e, d, 10)] == 1)
            else:
                model.Add(x[(e, d, 10)] == 0)

            # Apply other manual fixes (0-8)
            s_idx = fixed_indices.get((e, d))
            if s_idx is not None and s_idx < 9:
                model.Add(x[(e, d, s_idx)] == 1)

    # 5. Continuous work limit & Initial Streak
    for e in range(num_staff):
        if is_full_absent[e]: continue # Ver 8.23
        staff = staff_list[e]
        initial_streak = 0
        if staff.name.strip() in prev_shifts_map:
            for s in reversed(prev_shifts_map[staff.name.strip()]):
                if s in shift_names[1:8]: initial_streak += 1
                else: break
        
        for d in range(num_days - 5): 
            # Streak limit: indices 1-7 are work. Breaks are 0, 8, 9, 10
            model.Add(sum(x[(e, d+k, s)] for k in range(6) for s in range(1, 8)) <= 5)
        
        if initial_streak > 0:
            for k in range(1, 6):
                if initial_streak + k > 5:
                    # Must take a break (0, 8, 9, 10) somewhere in the first k days
                    model.Add(sum(x[(e, d_idx, s)] for d_idx in range(k) for s in [0, 8, 9, 10]) >= 1)

    # 6. Off days (8-12 days per person)
    for e in range(num_staff):
        if is_full_absent[e]: continue # Ver 8.23
        staff = staff_list[e]
        op = off_point_vars[e]
        is_long = long_term_absent[e]
        
        # Lower Limit: Still need 7 days (14 pts) of total non-work (including sick)
        # Rule (Ver 8.21): Relax lower limit for long-term absent staff to avoid artificial errors
        if not is_long:
            model.Add(op >= 14) 
        
        # Upper Limit: Normal off days + Ake days <= 13 days (26 pts).
        # We EXCLUDE Sick (9) and Absent (10) from the upper limit count.
        regulated_op = sum(x[(e, d, 0)] * 2 for d in range(num_days)) + \
                       sum(x[(e, d, 8)] * 1 for d in range(num_days))
        model.Add(regulated_op <= 26)

        # Rule: If previous month < 10 days, current month MUST BE >= 10 days (20 points)
        # We skip this for weekday-only staff (leaders/business) as their schedule is fixed
        if not staff.is_weekday_only:
            prev_count = prev_holiday_counts.get(staff.name.strip(), 10.0)
            if prev_count < 10.0:
                # Rule (Ver 8.21): Relax correction for long-term absent staff
                if is_long:
                    pass
                else:
                    # Soften this to high priority rather than hard constraint
                    recovery_met = model.NewBoolVar(f'e{e}_recovery_met')
                    model.Add(op >= 20).OnlyEnforceIf(recovery_met)
                    is_target_holiday.append(recovery_met) # Add to objective
            
    # 7. Night Shift Limit (Ver 8.23: SKIP FULL ABSENT)
    for e in range(num_staff):
        if is_full_absent[e]: continue
        model.Add(sum(x[(e, d, 6)] + x[(e, d, 7)] for d in range(num_days)) <= 8)

    # 8. 36 Agreement (Work Hour Limit: Max 176h) (Ver 8.23: SKIP FULL ABSENT)
    for e in range(num_staff):
        if is_full_absent[e]: continue
        total_hours = sum(x[(e, d, s)] * 8 for d in range(num_days) for s in range(1, 6)) + \
                      sum(x[(e, d, s)] * 12 for d in range(num_days) for s in range(6, 8))
        model.Add(total_hours <= 184)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    
    # Enable robust randomization for varied results
    import random
    seed_val = random.randint(1, 1000000)
    solver.parameters.random_seed = seed_val
    
    # Use multiple workers and randomize search for variety.
    # PORTFOLIO_SEARCH often gives better variety than DEFAULT_SEARCH.
    solver.parameters.num_search_workers = 8
    solver.parameters.randomize_search = True
    solver.parameters.search_branching = cp_model.PORTFOLIO_SEARCH
    
    

    # --- 8.5 Diagnostic: Pre-check Hard Conflicts (Ver 8.22, 8.23) ---
    for e, staff in enumerate(staff_list):
        if is_full_absent[e]: continue # Ver 8.23
        for d in range(num_days):
            s_idx = fixed_indices.get((e, d))
            if s_idx in [6, 7]: # Fixed Night
                # Check Next Day
                if d + 1 < num_days:
                    next_idx = fixed_indices.get((e, d+1))
                    if next_idx is not None and next_idx not in [0, 8, 9, 10]:
                        msg = f"衝突診断: {staff.name}様の{d+1}日の入力が不整合です。\n" \
                              f"前日({d}日)が「夜勤」のため、翌日は「明け」または「休み」である必要がありますが、" \
                              f"現在は「{shift_names[next_idx]}」が固定されています。"
                        return {"status": "error", "message": msg}
                # Check Next Next Day
                if d + 2 < num_days:
                    next2_idx = fixed_indices.get((e, d+2))
                    if next2_idx is not None and next2_idx not in [0, 9, 10]:
                        msg = f"衝突診断: {staff.name}様の{d+2}日の入力が不整合です。\n" \
                              f"前々日({d}日)が「夜勤」のため、この日は「休み」である必要がありますが、" \
                              f"現在は「{shift_names[next2_idx]}」が固定されています。"
                        return {"status": "error", "message": msg}
        
        # Check Month Boundary Conflicts
        staff_name = staff.name.strip()
        if staff_name in prev_shifts_map:
            p_shifts = prev_shifts_map[staff_name]
            if p_shifts and p_shifts[-1] in ["夜①", "夜②"]:
                d0_idx = fixed_indices.get((e, 0))
                if d0_idx is not None and d0_idx not in [0, 8, 9, 10]:
                    msg = f"衝突診断: {staff.name}様の1日の入力が不整合です。\n" \
                          f"前月末が「夜勤」のため、月初の1日は「明け」または「休み」である必要がありますが、" \
                          f"現在は「{shift_names[d0_idx]}」が固定されています。"
                    return {"status": "error", "message": msg}

    status = solver.Solve(model)
    
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        schedule_data = []
        for e, staff in enumerate(staff_list):
            staff_schedule = []
            for d in range(num_days):
                for s in range(num_shifts):
                    if solver.Value(x[(e, d, s)]) == 1:
                        staff_schedule.append(shift_names[s])
            schedule_data.append({"staff_id": staff.name, "floor": staff.floor, "shifts": staff_schedule})

        # Identify "holes" (undermanned mandatory shifts)
        holes = []
        for d, floor, s_idx, v in mandatory_shift_vars:
            if solver.Value(v) == 0:
                holes.append({
                    "day": d + 1,
                    "floor": floor,
                    "shift": shift_names[s_idx],
                    "count": 1 
                })

        res_status = "provisional" if len(holes) > 0 else "success"
        return {
            "status": res_status, 
            "schedule": schedule_data, 
            "num_days": num_days,
            "holes": holes,
            "has_shortage": len(holes) > 0,
            "message": "作成に成功しました(暫定案)" if len(holes) > 0 else "作成に成功しました"
        }
    else:
        msg = "条件が厳しすぎます。夜勤人数や労働制限を調整してください。"
        total_sick_days = sum(1 for (e, d), idx in fixed_indices.items() if idx in [9, 10])
        if total_sick_days > 20:
            msg += f"\n(現在、計{total_sick_days}日の病休・欠勤が入力されています。人員不足の可能性があります。)"
        return {"status": "error", "message": msg}

# ---- API Endpoints ----
@app.post("/api/generate")
def api_generate_schedule(req: ScheduleRequest, username: str = Depends(authenticate)):
    result = generate_schedule(req.year, req.month, req.staff_list, req.previous_shifts, req.previous_holiday_counts, req.current_shifts)
    if result["status"] == "failed":
        raise HTTPException(status_code=400, detail=result["message"])
    return result

class ExportRequest(BaseModel):
    schedule: List[dict]
    year: int
    month: int
    num_days: int = 30

@app.post("/api/export")
def api_export_schedule(req: ExportRequest, username: str = Depends(authenticate)):
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        num_cols = req.num_days + 9
        
        # --- Title Row ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"【{req.year}年 {req.month}月】 勤務予定表"
        title_cell.font = Font(name='Meiryo UI', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # --- Header Row ---
        weekdays_ja = ['日', '月', '火', '水', '木', '金', '土']
        headers = ["スタッフ", "階"]
        for d in range(req.num_days):
            dt = datetime.date(req.year, req.month, d + 1)
            wd = weekdays_ja[dt.weekday()]
            headers.append(f"{d+1}({wd})")
        headers.extend(["合計", "公休", "早②", "早③", "日", "遅①", "遅②"])
        
        header_font = Font(name='Meiryo UI', size=10, bold=True)
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        
        for col_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = h_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            
            # Weekend Highlighting
            if 2 < col_idx <= req.num_days + 2:
                d_num = col_idx - 2
                dt = datetime.date(req.year, req.month, d_num)
                if dt.weekday() == 5: # Sat
                    cell.fill = PatternFill(start_color="EBF5FF", end_color="EBF5FF", fill_type="solid")
                elif dt.weekday() == 6: # Sun
                    cell.fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        # --- Data Rows ---
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        standard_font = Font(name='Meiryo UI', size=10)
        
        # Sort schedule: floor 1, floor 2, floor 3
        sorted_schedule = sorted(req.schedule, key=lambda x: (x['floor'], x['staff_id']))
        
        current_row = 3
        for s in sorted_schedule:
            # Calculate metrics
            total_h = 0
            total_kyu = 0
            counts = {"早②": 0, "早③": 0, "日": 0, "遅①": 0, "遅②": 0}
            for shift in s["shifts"]:
                if shift in ["早②", "早③", "日", "遅①", "遅②"]: total_h += 8
                elif shift in ["夜①", "夜②"]: total_h += 12
                if shift == "休": total_kyu += 1.0
                elif shift == "明け": total_kyu += 0.5
                if shift in counts: counts[shift] += 1
            
            floor_label = f"{s['floor']}階"
            if s['floor'] == 3: floor_label = "業務員"
            
            row_data = [s['staff_id'], floor_label] + s['shifts'] + \
                       [f"{total_h}h", f"{total_kyu}日", counts["早②"], counts["早③"], counts["日"], counts["遅①"], counts["遅②"]]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = standard_font
                cell.border = border
                if col_idx > 1: cell.alignment = align_center
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # --- Column Widths ---
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 8
        for d in range(req.num_days):
            ws.column_dimensions[get_column_letter(d + 3)].width = 7.0
        
        stat_start = req.num_days + 3
        ws.column_dimensions[get_column_letter(stat_start)].width = 10
        ws.column_dimensions[get_column_letter(stat_start+1)].width = 8
        for i in range(5):
            ws.column_dimensions[get_column_letter(stat_start + 2 + i)].width = 6.0

        # --- Print Setup ---
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 8 # A3
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        
        file_id = str(uuid.uuid4())
        debug_fn = f"schedules/export_{file_id}.xlsx"
        with open(debug_fn, "wb") as f:
            f.write(content)
        
        return {"file_id": file_id}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/download/{file_id}")
def api_download_file(file_id: str, username: str = Depends(authenticate)):
    filename = f"schedules/export_{file_id}.xlsx"
    if not os.path.exists(filename):
        raise HTTPException(status_code=404, detail="File not found")
    
    # Generate a nice-looking filename for the user
    now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    display_name = f"shift_schedule_{now}.xlsx"
    
    return FileResponse(
        path=filename,
        filename=display_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

class SaveRequest(BaseModel):
    year: int
    month: int
    schedule: List[dict]

@app.post("/api/save")
def api_save_schedule(req: SaveRequest, username: str = Depends(authenticate)):
    try:
        filename = f"schedules/{req.year}_{req.month}.json"
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(req.schedule, f, ensure_ascii=False, indent=2)
        return {"status": "success", "message": "保存しました。"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/load")
def api_load_schedule(year: int, month: int, username: str = Depends(authenticate)):
    filename = f"schedules/{year}_{month}.json"
    if not os.path.exists(filename): return {"status": "not_found", "schedule": None}
    try:
        with open(filename, "r", encoding="utf-8") as f: schedule = json.load(f)
        return {"status": "success", "schedule": schedule}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
def read_root(username: str = Depends(authenticate)):
    html_path = os.path.join("static", "index_v8.html")
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()
    headers = {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0"
    }
    return HTMLResponse(content=content, headers=headers)

app.mount("/", StaticFiles(directory="static"), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
